import pandas as pd
import openpyxl
import logging
from time import time

# Set up logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

def vlookup_from_json(config):
    start_time = time()
    logger.info("Starting vlookup_from_json function")

    # Load the input file with 'openpyxl' to preserve formatting
    logger.info(f"Loading input file: {config['input_file']}")
    input_df = pd.read_excel(config['input_file'], engine='openpyxl')
    logger.info(f"Input file loaded. Shape: {input_df.shape}")

    # Pre-load all lookup files
    lookup_dfs = {}
    for lookup_key, lookup_config in config.items():
        if lookup_key.startswith('lookup_'):
            logger.info(f"Loading lookup file: {lookup_config['lookup_file']}")
            lookup_dfs[lookup_key] = pd.read_excel(lookup_config['lookup_file'], engine='openpyxl')
            logger.info(f"Lookup file {lookup_key} loaded. Shape: {lookup_dfs[lookup_key].shape}")

    # Optimize lookup function
    def perform_lookup(input_values, lookup_config, lookup_df):
        column_a = lookup_config['lookup_column_a']
        column_b = lookup_config['lookup_column_b']
        
        # Create sets for faster lookups
        set_a = set(lookup_df[column_a])
        set_b = set(lookup_df[column_b])
        
        def lookup_logic(value):
            # If input value is empty, return an empty string
            if pd.isna(value) or value == "":
                return ""

            # Perform lookup based on the value presence in column_a and column_b
            in_a = value in set_a
            in_b = value in set_b
            if in_a and in_b:
                return lookup_config['both_columns']
            elif in_a:
                return lookup_config['only_column_a']
            elif in_b:
                return lookup_config['only_column_b']
            else:
                return lookup_config['na_value']
        
        # Vectorized lookup
        results = input_values.map(lookup_logic)
        return results

    # Apply lookups and store in a temporary dataframe
    for lookup_key, lookup_config in config.items():
        if lookup_key.startswith('lookup_'):
            logger.info(f"Performing lookup: {lookup_key}")
            start_lookup = time()
            
            # Add new column with the lookup results to a temp dataframe
            input_df[lookup_key] = perform_lookup(
                input_df[config['input_lookup_column']],
                lookup_config,
                lookup_dfs[lookup_key]
            )
            
            logger.info(f"Lookup {lookup_key} completed in {time() - start_lookup:.2f} seconds")

    # Load the workbook and select the active sheet to preserve formatting
    workbook = openpyxl.load_workbook(config['input_file'])
    sheet = workbook.active

    # Find the column index for the specified output column
    output_column = config.get('output_column', 'NewColumn')  # Default to 'NewColumn' if not provided
    output_column_idx = None
    
    for col in sheet.iter_cols(1, sheet.max_column, 1, 1):
        if col[0].value == output_column:
            output_column_idx = col[0].column  # Get column index

    if output_column_idx is None:
        # If output_column doesn't exist, append at the end of the sheet
        output_column_idx = sheet.max_column + 1
        sheet.cell(row=1, column=output_column_idx).value = output_column  # Add new header

    # Write values to the specified output column
    for row_idx, value in enumerate(input_df[lookup_key], start=2):  # Skip header
        sheet.cell(row=row_idx, column=output_column_idx).value = value

    # Save the updated workbook
    output_file = "updated_" + config['input_file']
    workbook.save(output_file)
    logger.info(f"Updated file saved to {output_file}")

    total_time = time() - start_time
    logger.info(f"vlookup_from_json completed in {total_time:.2f} seconds")

# Example config from the JSON-like structure
config = {
    "input_file": "input_data.xlsx",
    "input_lookup_column": "C",
    "output_column": "XYZ",  # Specify the column where you want to insert the new data
    "lookup_1": {
        "lookup_file": "lookup1.xlsx",
        "lookup_column_a": "A",
        "lookup_column_b": "B",
        "both_columns": "xyz",
        "only_column_a": "ABC",
        "only_column_b": "ZZZ",
        "na_value": "YYY"
    },
}

if __name__ == "__main__":
    # Call the function with the provided config
    vlookup_from_json(config)
