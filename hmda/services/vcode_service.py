import json
import os
import pandas as pd
from typing import Any, List, Dict
from hmda.utils.logger import get_logger
from openpyxl import load_workbook

logger = get_logger(__name__)

class VCodeService:

    @staticmethod
    def execute(lite_auto_file: str, file2: str, file3: str, file4: str):
        VCodeService.handle_file1(lite_auto_file)
        VCodeService.handle_file2(file2)
        VCodeService.handle_file3(file3)
        VCodeService.handle_file4(file4)

    @staticmethod
    def handle_file1(file1: str):
        config_path = os.path.join(os.path.dirname(__file__), '..', 'config', 'vcode.json')
        lite_auto_config = VCodeService.load_vcode(config_path)

        VCodeService.replace_vcode(file1, lite_auto_config)

    @staticmethod
    def handle_file2(file2: str):
        config_path = os.path.join(os.path.dirname(__file__), '..', 'config', 'vcode.json')
        file2_config = VCodeService.load_vcode(config_path)

        VCodeService.replace_vcode(file2, file2_config)
    
    @staticmethod
    def handle_file3(file3: str):
        config_path = os.path.join(os.path.dirname(__file__), '..', 'config', 'vcode.json')
        file3_config = VCodeService.load_vcode(config_path)

        VCodeService.replace_vcode(file3, file3_config)

    @staticmethod
    def handle_file4(file4: str):
        config_path = os.path.join(os.path.dirname(__file__), '..', 'config', 'vcode.json')
        file4_config = VCodeService.load_vcode(config_path)

        VCodeService.replace_vcode(file4, file4_config)

    @staticmethod
    def replace_vcode(excel_file: str, config: Dict[str, Any]):
        logger.info(f"Starting file replacement process for {excel_file}")
        
        # Load the workbook and select the active worksheet
        workbook = load_workbook(excel_file)

        sheet_name = config.get('sheet_name', 'Sheet1')
        edit_column = config['edit_column']
        replacers = config['replacers']
        replaced_count = 0

        if sheet_name == 'last-sheet':
            worksheet = workbook.worksheets[-1]
        elif sheet_name == 'first-sheet':
            worksheet = workbook.worksheets[0]
        elif sheet_name == 'active-sheet':
            worksheet = workbook.active
        else:
            worksheet = workbook[sheet_name]
        
        # Find the column index for the edit_column
        header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
        try:
            edit_column_index = header_row.index(edit_column) + 1
        except ValueError:
            logger.warning(f"Edit column '{edit_column}' not found in the Excel file '{excel_file}'.")
            return

        # Apply vcode-mapping to the specific edit column
        for row in worksheet.iter_rows(min_row=2, min_col=edit_column_index, max_col=edit_column_index):
            cell = row[0]
            old_value = str(cell.value).strip() 
            new_value = replacers.get(old_value)
            if new_value is not None and new_value != old_value:
                cell.value = new_value
                replaced_count += 1

        # Save the changes back to the same file
        workbook.save(excel_file)

        logger.info(f"Vcode replacement process completed. Replaced {replaced_count} values in column '{edit_column}'.")
        logger.info(f"Updated data saved to {excel_file}")

    @staticmethod
    def load_vcode(json_file: str) -> Dict[str, Any]:
        with open(json_file, 'r') as f:
            config = json.load(f)
        return {
            'edit_column': config['edit_column'],
            'sheet_name': config['sheet_name'],
            'replacers': config['replacers']
        }
    
